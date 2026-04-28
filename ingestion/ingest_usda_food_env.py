"""
Ingests USDA Food Environment Atlas data at county level.

Source: USDA Economic Research Service
Download: https://www.ers.usda.gov/media/5569/food-environment-atlas-data-download.xlsx

The file is an Excel workbook with multiple sheets. This script reads
the relevant sheets and loads selected metrics into usda_food_env.

Metrics loaded:
  Sheet "ACCESS"         → PCT_LACCESS_POP15  → pct_low_food_access
  Sheet "STORES"         → GROCPTH16          → groceries_per_1000
  Sheet "RESTAURANTS"    → FFRPTH16           → fast_food_per_1000
  Sheet "ASSISTANCE"     → PCT_SNAP17         → pct_snap
  Sheet "LOCAL"          → FMRKT18            → farmers_markets

The FIPS column in each sheet is named "FIPS" and contains 5-digit county codes.

Usage:
  python ingest_usda_food_env.py --file /path/to/FoodEnvironmentAtlas.xls [--data-year 2018]
  python ingest_usda_food_env.py --download [--data-year 2018]
"""

import argparse
import io
import logging
import os

import psycopg2
import requests
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "../config/.env"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
log = logging.getLogger(__name__)

USDA_URL = (
    "https://www.ers.usda.gov/media/5569/food-environment-atlas-data-download.xlsx"
)

# (sheet_name, source_column, dest_column, numeric_type)
COLUMNS = [
    ("ACCESS",      "PCT_LACCESS_POP15", "pct_low_food_access", float),
    ("STORES",      "GROCPTH16",         "groceries_per_1000",  float),
    ("RESTAURANTS", "FFRPTH16",          "fast_food_per_1000",  float),
    ("ASSISTANCE",  "PCT_SNAP17",        "pct_snap",            float),
    ("LOCAL",       "FMRKT18",           "farmers_markets",     int),
]


def _safe(value, cast):
    if value is None or value == "" or str(value).strip() in ("", "NA", "N/A", "-"):
        return None
    try:
        return cast(value)
    except (ValueError, TypeError):
        return None


def load_workbook_data(path_or_bytes) -> dict[str, dict]:
    """Return {fips: {col: value}} from the Excel file."""
    import openpyxl

    if isinstance(path_or_bytes, (str, os.PathLike)):
        wb = openpyxl.load_workbook(path_or_bytes, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)

    records: dict[str, dict] = {}

    for sheet_name, src_col, dst_col, cast in COLUMNS:
        if sheet_name not in wb.sheetnames:
            log.warning("Sheet '%s' not found in workbook — skipping %s", sheet_name, dst_col)
            continue

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        next(rows)  # skip title row (row 1 is the sheet name, row 2 is the header)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]

        fips_idx = next((i for i, h in enumerate(header) if h.upper() == "FIPS"), None)
        col_idx  = next((i for i, h in enumerate(header) if h.upper() == src_col.upper()), None)

        if fips_idx is None or col_idx is None:
            log.warning("Could not find FIPS or %s in sheet '%s'", src_col, sheet_name)
            continue

        for row in rows:
            if row[fips_idx] is None:
                continue
            try:
                fips = str(int(row[fips_idx])).zfill(5)
            except (ValueError, TypeError):
                continue  # skip N/A or non-numeric FIPS (state totals, etc.)
            val  = _safe(row[col_idx], cast)
            records.setdefault(fips, {})[dst_col] = val

    return records


def upsert(conn, records: dict[str, dict], known_fips: set[str], data_year: int) -> int:
    sql = """
        INSERT INTO usda_food_env
            (fips, data_year, pct_low_food_access, groceries_per_1000,
             fast_food_per_1000, pct_snap, farmers_markets)
        VALUES
            (%(fips)s, %(data_year)s, %(pct_low_food_access)s, %(groceries_per_1000)s,
             %(fast_food_per_1000)s, %(pct_snap)s, %(farmers_markets)s)
        ON CONFLICT (fips, data_year) DO UPDATE SET
            pct_low_food_access = EXCLUDED.pct_low_food_access,
            groceries_per_1000  = EXCLUDED.groceries_per_1000,
            fast_food_per_1000  = EXCLUDED.fast_food_per_1000,
            pct_snap            = EXCLUDED.pct_snap,
            farmers_markets     = EXCLUDED.farmers_markets,
            fetched_at          = NOW()
    """
    count = 0
    with conn.cursor() as cur:
        for fips, data in records.items():
            if fips not in known_fips:
                continue
            row = {
                "fips":               fips,
                "data_year":          data_year,
                "pct_low_food_access": data.get("pct_low_food_access"),
                "groceries_per_1000":  data.get("groceries_per_1000"),
                "fast_food_per_1000":  data.get("fast_food_per_1000"),
                "pct_snap":            data.get("pct_snap"),
                "farmers_markets":     data.get("farmers_markets"),
            }
            cur.execute(sql, row)
            count += cur.rowcount
    conn.commit()
    return count


def ingest(conn, file_path: str | None, data_year: int, download: bool = False) -> int:
    with conn.cursor() as cur:
        cur.execute("SELECT fips FROM geo_entities WHERE geo_type = 'county'")
        known_fips = {r[0] for r in cur.fetchall()}

    if download or file_path is None:
        log.info("Downloading USDA Food Environment Atlas from %s ...", USDA_URL)
        resp = requests.get(USDA_URL, timeout=120)
        resp.raise_for_status()
        source = resp.content
    else:
        source = file_path

    records = load_workbook_data(source)
    log.info("Loaded %d county records from workbook", len(records))

    count = upsert(conn, records, known_fips, data_year)
    log.info("Upserted %d rows into usda_food_env", count)
    return count


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file",     help="Path to FoodEnvironmentAtlas.xls")
    group.add_argument("--download", action="store_true",
                       help="Download the file from USDA ERS automatically")
    parser.add_argument("--data-year", type=int, default=2018,
                        help="Vintage year to tag this dataset (default: 2018)")
    args = parser.parse_args()

    conn = psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=os.environ["DB_PORT"],
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ.get("DB_PASSWORD", ""),
    )
    try:
        ingest(conn, args.file, args.data_year, args.download)
        log.info("Done.")
    finally:
        conn.close()
