"""
Tests for ingest_cdc_places, ingest_bls_laus, ingest_usda_food_env, and ingest_epa_aqi.
All DB and HTTP calls are mocked.
"""
import io
import json
from collections import defaultdict
from unittest.mock import MagicMock, patch, call

import pytest


# ---------------------------------------------------------------------------
# CDC PLACES
# ---------------------------------------------------------------------------

from ingest_cdc_places import pivot, fetch_places, upsert as cdc_upsert, ingest as cdc_ingest

SAMPLE_PLACES_ROWS = [
    {"locationid": "06037", "year": "2022", "measureid": "OBESITY",    "data_value": "36.5"},
    {"locationid": "06037", "year": "2022", "measureid": "DIABETES",   "data_value": "10.2"},
    {"locationid": "06037", "year": "2022", "measureid": "CSMOKING",   "data_value": "12.0"},
    {"locationid": "06037", "year": "2022", "measureid": "BPHIGH",     "data_value": "38.1"},
    {"locationid": "06037", "year": "2022", "measureid": "DEPRESSION", "data_value": "22.4"},
    {"locationid": "06037", "year": "2022", "measureid": "LPA",        "data_value": "30.1"},
    {"locationid": "06037", "year": "2022", "measureid": "MHLTH",      "data_value": "15.6"},
    {"locationid": "48201", "year": "2022", "measureid": "OBESITY",    "data_value": "38.0"},
]


class TestCdcPivot:
    def test_known_measures_mapped(self):
        records = pivot(SAMPLE_PLACES_ROWS, 2022)
        assert "06037" in records
        assert records["06037"]["pct_obesity"] == 36.5
        assert records["06037"]["pct_diabetes"] == 10.2
        assert records["06037"]["pct_smoking"] == 12.0

    def test_year_set(self):
        records = pivot(SAMPLE_PLACES_ROWS, 2022)
        assert records["06037"]["year"] == 2022

    def test_fips_zero_padded(self):
        rows = [{"locationid": "6037", "year": "2022", "measureid": "OBESITY", "data_value": "36.5"}]
        records = pivot(rows, 2022)
        assert "06037" in records

    def test_unknown_measure_ignored(self):
        rows = [{"locationid": "06037", "year": "2022", "measureid": "UNKNOWN", "data_value": "9.9"}]
        records = pivot(rows, 2022)
        assert "06037" not in records or "UNKNOWN" not in records.get("06037", {})

    def test_null_value_skipped(self):
        rows = [{"locationid": "06037", "year": "2022", "measureid": "OBESITY", "data_value": None}]
        records = pivot(rows, 2022)
        assert records["06037"].get("pct_obesity") is None

    def test_multiple_fips(self):
        records = pivot(SAMPLE_PLACES_ROWS, 2022)
        assert "48201" in records
        assert records["48201"]["pct_obesity"] == 38.0


class TestCdcFetch:
    def _mock_resp(self, data):
        m = MagicMock()
        m.json.return_value = data
        m.raise_for_status = MagicMock()
        return m

    def test_single_page(self):
        with patch("ingest_cdc_places.requests.get",
                   return_value=self._mock_resp(SAMPLE_PLACES_ROWS[:3])):
            rows = fetch_places(2022)
        assert len(rows) == 3

    def test_paginate_until_short_page(self):
        # first page full (50000), second page short
        big_page = SAMPLE_PLACES_ROWS * 2  # only 16 rows — less than PAGE_LIMIT of 50000
        with patch("ingest_cdc_places.requests.get",
                   return_value=self._mock_resp(big_page)):
            rows = fetch_places(2022)
        assert len(rows) == len(big_page)


class TestCdcUpsert:
    def _make_conn(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.rowcount = 1
        return mock_conn, mock_cur

    def test_skips_unknown_fips(self):
        conn, cur = self._make_conn()
        records = {"99999": {"year": 2022, "pct_obesity": 30.0}}
        count = cdc_upsert(conn, records, known_fips={"06037"})
        cur.execute.assert_not_called()

    def test_upserts_known_fips(self):
        conn, cur = self._make_conn()
        records = pivot(SAMPLE_PLACES_ROWS, 2022)
        cdc_upsert(conn, records, known_fips={"06037", "48201"})
        assert cur.execute.call_count == 2

    def test_commits_after_upsert(self):
        conn, cur = self._make_conn()
        records = pivot(SAMPLE_PLACES_ROWS[:1], 2022)
        cdc_upsert(conn, records, known_fips={"06037"})
        conn.commit.assert_called_once()


class TestCdcIngest:
    def test_ingest_calls_fetch_and_upsert(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.fetchall.return_value = [("06037",), ("48201",)]
        mock_cur.rowcount = 1

        with patch("ingest_cdc_places.fetch_places", return_value=SAMPLE_PLACES_ROWS) as mock_fetch:
            count = cdc_ingest(mock_conn, 2022)

        mock_fetch.assert_called_once_with(2022, None)
        assert count >= 0


# ---------------------------------------------------------------------------
# BLS LAUS
# ---------------------------------------------------------------------------

from ingest_bls_laus import (
    parse_flat_file,
    upsert as bls_upsert,
    ingest as bls_ingest,
)

_SAMPLE_FLAT = (
    "LAUS Code\tState FIPS\tCounty FIPS\tArea Title\tYear\tPeriod\t"
    "Labor Force\tEmployed\tUnemployed\tRate\n"
    "CN0603700000000\t06\t037\tLos Angeles County, CA\t2022\tAnnual\t"
    "5,000,000\t4,750,000\t250,000\t5.0\n"
    "CN0100100000000\t01\t001\tAutauga County, AL\t2022\tAnnual\t"
    "28,617\t27,978\t639\t2.2\n"
)


class TestBlsParseFlatFile:
    def test_basic_row(self):
        records = parse_flat_file(_SAMPLE_FLAT)
        assert ("06037", 2022) in records

    def test_fips_constructed(self):
        records = parse_flat_file(_SAMPLE_FLAT)
        assert ("01001", 2022) in records

    def test_comma_stripped_from_labor_force(self):
        records = parse_flat_file(_SAMPLE_FLAT)
        assert records[("06037", 2022)]["labor_force"] == 5_000_000

    def test_unemployment_rate_float(self):
        records = parse_flat_file(_SAMPLE_FLAT)
        assert records[("06037", 2022)]["unemployment_rate"] == 5.0

    def test_header_row_skipped(self):
        records = parse_flat_file(_SAMPLE_FLAT)
        assert len(records) == 2

    def test_bytes_input(self):
        records = parse_flat_file(_SAMPLE_FLAT.encode())
        assert ("06037", 2022) in records

    def test_non_county_rows_skipped(self):
        content = "Some header line\nAnother line\n" + _SAMPLE_FLAT
        records = parse_flat_file(content)
        assert len(records) == 2

    def test_missing_value_returns_none(self):
        line = "CN0600100000000\t06\t001\tAlpine County, CA\t2022\tAnnual\t\t\t\t\n"
        records = parse_flat_file(line)
        assert records[("06001", 2022)]["labor_force"] is None
        assert records[("06001", 2022)]["unemployment_rate"] is None


class TestBlsUpsert:
    def _make_conn(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.rowcount = 1
        return mock_conn, mock_cur

    def test_upsert_executes_once_per_record(self):
        conn, cur = self._make_conn()
        records = {("06037", 2022): {"unemployment_rate": 5.0, "labor_force": 5000000}}
        bls_upsert(conn, records)
        assert cur.execute.call_count == 1

    def test_commits(self):
        conn, cur = self._make_conn()
        bls_upsert(conn, {})
        conn.commit.assert_called_once()


def _bls_api_response(fips_list, start, end):
    """Build a fake BLS API v2 response for the given counties."""
    series = []
    for fips in fips_list:
        for measure, col in {"06": "labor_force", "05": "employed",
                             "04": "unemployed",  "03": "unemployment_rate"}.items():
            sid = f"LAUCN{fips}0000000{measure}"
            data = [{"year": str(y), "period": "M13", "value": "1000"} for y in range(start, end + 1)]
            series.append({"seriesID": sid, "data": data})
    return {"status": "REQUEST_SUCCEEDED", "Results": {"series": series}}


class TestBlsIngest:
    def _make_conn(self, fips_list=("06037",)):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_cur.fetchall.return_value = [(f,) for f in fips_list]
        mock_cur.rowcount = 1
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        return mock_conn, mock_cur

    def test_ingest_posts_to_bls_api(self):
        conn, _ = self._make_conn()
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = _bls_api_response(["06037"], 2021, 2022)
        with patch("ingest_bls_laus.requests.post", return_value=mock_resp) as mock_post:
            bls_ingest(conn, 2021, 2022)
        assert mock_post.called
        assert mock_post.call_args[0][0] == "https://api.bls.gov/publicAPI/v2/timeseries/data/"

    def test_ingest_includes_years_in_payload(self):
        conn, _ = self._make_conn()
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = _bls_api_response(["06037"], 2022, 2022)
        with patch("ingest_bls_laus.requests.post", return_value=mock_resp) as mock_post:
            bls_ingest(conn, 2022, 2022)
        payload = mock_post.call_args[1]["json"]
        assert payload["startyear"] == "2022"
        assert payload["endyear"] == "2022"


# ---------------------------------------------------------------------------
# USDA Food Environment Atlas
# ---------------------------------------------------------------------------

from ingest_usda_food_env import load_workbook_data, upsert as usda_upsert, _safe


def _make_workbook(sheets: dict[str, list[list]]):
    """Build a minimal openpyxl workbook in memory for testing.

    Prepends a dummy title row to each sheet so that row 1 is the title (skipped
    by ingestion) and row 2 is the actual header, matching the real USDA xlsx layout.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append([sheet_name])  # title row — ingestion skips this
        for row in rows:
            ws.append(row)
    return wb


class TestUsdaSafe:
    def test_float_conversion(self):
        assert _safe("12.5", float) == 12.5

    def test_int_conversion(self):
        assert _safe("45", int) == 45

    def test_na_returns_none(self):
        assert _safe("NA", float) is None

    def test_empty_returns_none(self):
        assert _safe("", float) is None

    def test_none_returns_none(self):
        assert _safe(None, float) is None


class TestUsdaLoadWorkbook:
    def _workbook_bytes(self, sheets):
        import io, openpyxl
        wb = _make_workbook(sheets)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_loads_pct_low_food_access(self):
        data = self._workbook_bytes({
            "ACCESS": [["FIPS", "PCT_LACCESS_POP15"], [6037, 12.5]],
        })
        records = load_workbook_data(data)
        assert "06037" in records
        assert records["06037"]["pct_low_food_access"] == 12.5

    def test_fips_zero_padded(self):
        data = self._workbook_bytes({
            "STORES": [["FIPS", "GROCPTH16"], [1001, 0.42]],
        })
        records = load_workbook_data(data)
        assert "01001" in records

    def test_missing_sheet_skipped(self):
        # Only ACCESS sheet — STORES not present
        data = self._workbook_bytes({
            "ACCESS": [["FIPS", "PCT_LACCESS_POP15"], [6037, 12.5]],
        })
        records = load_workbook_data(data)
        # No groceries_per_1000 since STORES sheet missing
        assert "groceries_per_1000" not in records.get("06037", {})

    def test_na_value_stored_as_none(self):
        data = self._workbook_bytes({
            "LOCAL": [["FIPS", "FMRKT18"], [6037, None]],
        })
        records = load_workbook_data(data)
        assert records.get("06037", {}).get("farmers_markets") is None

    def test_multiple_sheets_merged(self):
        data = self._workbook_bytes({
            "ACCESS":      [["FIPS", "PCT_LACCESS_POP15"], [6037, 12.5]],
            "STORES":      [["FIPS", "GROCPTH16"],         [6037, 0.42]],
            "RESTAURANTS": [["FIPS", "FFRPTH16"],          [6037, 2.10]],
            "ASSISTANCE":  [["FIPS", "PCT_SNAP17"],        [6037, 14.2]],
            "LOCAL":       [["FIPS", "FMRKT18"],           [6037, 45]],
        })
        records = load_workbook_data(data)
        r = records["06037"]
        assert r["pct_low_food_access"] == 12.5
        assert r["groceries_per_1000"] == 0.42
        assert r["fast_food_per_1000"] == 2.10
        assert r["pct_snap"] == 14.2
        assert r["farmers_markets"] == 45


class TestUsdaUpsert:
    def _make_conn(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.rowcount = 1
        return mock_conn, mock_cur

    def test_skips_unknown_fips(self):
        conn, cur = self._make_conn()
        records = {"99999": {"pct_low_food_access": 10.0}}
        usda_upsert(conn, records, known_fips={"06037"}, data_year=2018)
        cur.execute.assert_not_called()

    def test_upserts_known_fips(self):
        conn, cur = self._make_conn()
        records = {"06037": {"pct_low_food_access": 12.5, "groceries_per_1000": 0.42,
                             "fast_food_per_1000": 2.10, "pct_snap": 14.2, "farmers_markets": 45}}
        usda_upsert(conn, records, known_fips={"06037"}, data_year=2018)
        cur.execute.assert_called_once()

    def test_commits(self):
        conn, cur = self._make_conn()
        usda_upsert(conn, {}, known_fips=set(), data_year=2018)
        conn.commit.assert_called_once()


# ---------------------------------------------------------------------------
# EPA AQI
# ---------------------------------------------------------------------------

from ingest_epa_aqi import (
    parse_aqi_csv, normalize_county, match_to_fips, upsert as epa_upsert,
)

_SAMPLE_AQI_CSV = (
    "State,County,Year,Days with AQI,Good Days,Moderate Days,"
    "Unhealthy for Sensitive Groups Days,Unhealthy Days,Very Unhealthy Days,"
    "Hazardous Days,Max AQI,90th Percentile AQI,Median AQI,"
    "Days CO,Days NO2,Days Ozone,Days PM2.5,Days PM10\n"
    "California,Los Angeles,2022,363,180,120,50,10,3,0,112,78,42,"
    "0,5,95,85,10\n"
    "Texas,Harris,2022,355,150,140,55,15,5,1,145,90,55,"
    "1,10,80,100,8\n"
)


class TestEpaAqiNormalizeCounty:
    def test_strips_county(self):
        assert normalize_county("Los Angeles County") == "los angeles"

    def test_strips_parish(self):
        assert normalize_county("St. Tammany Parish") == "st. tammany"

    def test_strips_borough(self):
        assert normalize_county("Fairbanks North Star Borough") == "fairbanks north star"

    def test_strips_city_and_borough(self):
        assert normalize_county("Juneau City and Borough") == "juneau"

    def test_already_normalized(self):
        assert normalize_county("Los Angeles") == "los angeles"

    def test_lowercases(self):
        assert normalize_county("HARRIS COUNTY") == "harris"


class TestEpaAqiParseCsv:
    def test_parses_two_rows(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV)
        assert len(rows) == 2

    def test_state_and_county(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV)
        assert rows[0]["state"] == "California"
        assert rows[0]["county"] == "Los Angeles"

    def test_numeric_fields(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV)
        r = rows[0]
        assert r["year"] == 2022
        assert r["good_days"] == 180
        assert r["median_aqi"] == 42.0
        assert r["max_aqi"] == 112
        assert r["pm25_days"] == 85
        assert r["ozone_days"] == 95

    def test_bytes_input(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV.encode())
        assert len(rows) == 2

    def test_empty_field_returns_none(self):
        csv_with_blank = (
            "State,County,Year,Days with AQI,Good Days,Moderate Days,"
            "Unhealthy for Sensitive Groups Days,Unhealthy Days,Very Unhealthy Days,"
            "Hazardous Days,Max AQI,90th Percentile AQI,Median AQI,"
            "Days CO,Days NO2,Days Ozone,Days PM2.5,Days PM10\n"
            "California,Alpine,2022,,,,,,,,,,,,,,\n"
        )
        rows = parse_aqi_csv(csv_with_blank)
        assert rows[0]["good_days"] is None
        assert rows[0]["median_aqi"] is None


class TestEpaAqiMatchToFips:
    def _geo_lookup(self):
        return {
            ("06", "los angeles"): "06037",
            ("48", "harris"): "48201",
        }

    def test_matches_known_counties(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV)
        records = match_to_fips(rows, self._geo_lookup())
        assert ("06037", 2022) in records
        assert ("48201", 2022) in records

    def test_unmatched_state_skipped(self):
        rows = [{"state": "Unknown State", "county": "Somewhere", "year": 2022,
                 "good_days": 100, "median_aqi": 40.0}]
        records = match_to_fips(rows, self._geo_lookup())
        assert len(records) == 0

    def test_unmatched_county_skipped(self):
        rows = [{"state": "California", "county": "Nonexistent County", "year": 2022,
                 "good_days": 100, "median_aqi": 40.0}]
        records = match_to_fips(rows, self._geo_lookup())
        assert len(records) == 0

    def test_metrics_in_record(self):
        rows = parse_aqi_csv(_SAMPLE_AQI_CSV)
        records = match_to_fips(rows, self._geo_lookup())
        r = records[("06037", 2022)]
        assert r["good_days"] == 180
        assert r["median_aqi"] == 42.0


class TestEpaAqiUpsert:
    def _make_conn(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.rowcount = 1
        return mock_conn, mock_cur

    def test_executes_once_per_record(self):
        conn, cur = self._make_conn()
        records = {("06037", 2022): {"days_with_aqi": 363, "good_days": 180,
                                     "moderate_days": 120, "unhealthy_sensitive_days": 50,
                                     "unhealthy_days": 10, "very_unhealthy_days": 3,
                                     "hazardous_days": 0, "max_aqi": 112, "median_aqi": 42.0,
                                     "pm25_days": 85, "ozone_days": 95}}
        epa_upsert(conn, records)
        assert cur.execute.call_count == 1

    def test_commits(self):
        conn, cur = self._make_conn()
        epa_upsert(conn, {})
        conn.commit.assert_called_once()


# ---------------------------------------------------------------------------
# FBI Crime
# ---------------------------------------------------------------------------

from ingest_fbi_crime import (
    parse_crime_csv, upsert as crime_upsert,
)

_SAMPLE_CRIME_CSV = (
    "State,State Code,County,County Code,Year,Population,Violent Crime,Property Crime\n"
    "California,06,Los Angeles,037,2022,10014009,15234,89023\n"
    "Texas,48,Harris,201,2022,4731145,23456,145678\n"
)

# Two agencies in the same county to test aggregation
_SAMPLE_CRIME_CSV_MULTI = (
    "State,State Code,County,County Code,Year,Population,Violent Crime,Property Crime\n"
    "California,06,Los Angeles,037,2022,6000000,10000,60000\n"
    "California,06,Los Angeles,037,2022,4014009,5234,29023\n"
)


class TestFbiCrimeParseCsv:
    def test_parses_two_counties(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV)
        assert len(records) == 2

    def test_fips_constructed(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV)
        assert ("06037", 2022) in records
        assert ("48201", 2022) in records

    def test_violent_crime_count(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV)
        assert records[("06037", 2022)]["violent_crimes"] == 15234

    def test_property_crime_count(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV)
        assert records[("48201", 2022)]["property_crimes"] == 145678

    def test_rate_computed_per_100k(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV)
        r = records[("06037", 2022)]
        expected = round(15234 / 10014009 * 100_000, 1)
        assert r["violent_crime_rate"] == expected

    def test_multiple_agencies_aggregated(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV_MULTI)
        assert len(records) == 1
        r = records[("06037", 2022)]
        assert r["violent_crimes"] == 15234      # 10000 + 5234
        assert r["population_covered"] == 10014009  # 6000000 + 4014009

    def test_bytes_input(self):
        records = parse_crime_csv(_SAMPLE_CRIME_CSV.encode())
        assert ("06037", 2022) in records

    def test_zero_population_rate_is_none(self):
        csv = (
            "State,State Code,County,County Code,Year,Population,Violent Crime,Property Crime\n"
            "California,06,Tiny,999,2022,0,0,0\n"
        )
        records = parse_crime_csv(csv)
        r = records[("06999", 2022)]
        assert r["violent_crime_rate"] is None
        assert r["property_crime_rate"] is None


class TestFbiCrimeUpsert:
    def _make_conn(self):
        mock_conn = MagicMock()
        mock_cur = MagicMock()
        mock_conn.cursor.return_value.__enter__ = MagicMock(return_value=mock_cur)
        mock_conn.cursor.return_value.__exit__ = MagicMock(return_value=False)
        mock_cur.rowcount = 1
        return mock_conn, mock_cur

    def test_executes_once_per_known_fips(self):
        conn, cur = self._make_conn()
        records = {("06037", 2022): {"population_covered": 10014009,
                                     "violent_crimes": 15234, "violent_crime_rate": 450.1,
                                     "property_crimes": 89023, "property_crime_rate": 2630.5}}
        crime_upsert(conn, records, known_fips={"06037"})
        assert cur.execute.call_count == 1

    def test_unknown_fips_skipped(self):
        conn, cur = self._make_conn()
        records = {("99999", 2022): {"population_covered": 1000,
                                     "violent_crimes": 5, "violent_crime_rate": 500.0,
                                     "property_crimes": 20, "property_crime_rate": 2000.0}}
        crime_upsert(conn, records, known_fips={"06037"})
        assert cur.execute.call_count == 0

    def test_commits(self):
        conn, cur = self._make_conn()
        crime_upsert(conn, {}, known_fips=set())
        conn.commit.assert_called_once()
